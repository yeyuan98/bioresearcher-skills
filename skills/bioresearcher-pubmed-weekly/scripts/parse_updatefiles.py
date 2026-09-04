#!/usr/bin/env python3
"""Streaming parser for PubMed updatefiles XML / XML.gz into one Excel workbook.

Reads one or more NCBI PubMed updatefiles (``pubmedNNnNNNN.xml`` or
``.xml.gz``) and produces a single ``.xlsx`` workbook:

- Sheet ``PubMed Articles``: columns PMID, DOI, Title, Journal, ISSN, PubDate,
  FirstAuthor, LastAuthor, PublicationTypes (exact frozen order).
- Sheet ``Deleted PMIDs``: single column PMID.

Updatefiles interleave ``<PubmedArticle>`` and ``<DeleteCitation>`` records;
both are handled. Parsing is memory-bounded: ``xml.etree.iterparse`` with
element clearing plus an ``openpyxl`` write_only workbook. Output ordering is
deterministic: input file order on the command line, document order within
each file. Depends only on the Python standard library and openpyxl.
"""

import argparse
import gzip
import json
import os
import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook

ARTICLE_SHEET = "PubMed Articles"
DELETED_SHEET = "Deleted PMIDs"

COLUMNS = [
    "PMID",
    "DOI",
    "Title",
    "Journal",
    "ISSN",
    "PubDate",
    "FirstAuthor",
    "LastAuthor",
    "PublicationTypes",
]

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10,
    "november": 11, "december": 12,
}

GZIP_MAGIC = b"\x1f\x8b"


def open_maybe_gzip(path):
    """Open ``path`` as a gzip stream if it carries the gzip magic bytes."""
    with open(path, "rb") as probe:
        magic = probe.read(2)
    if magic == GZIP_MAGIC:
        return gzip.open(path, "rb")
    return open(path, "rb")


def join_text(elem):
    """Return concatenated (markup-flattened) stripped text of ``elem``."""
    if elem is None:
        return ""
    return "".join(elem.itertext()).strip()


def normalize_month(value):
    """Map a Month element value (name or number) to a 1-12 int, or None."""
    if not value:
        return None
    text = value.strip()
    if text.isdigit():
        number = int(text)
        if 1 <= number <= 12:
            return number
        return None
    return MONTHS.get(text.lower())


def normalize_pubdate(pubdate):
    """Normalize JournalIssue/PubDate to ``YYYY-MM`` or ``YYYY`` (or ``""``).

    Accepts Year plus Month given as a name (``Jan`` / ``January``) or a
    number (``1`` / ``01``). Day is consumed by PubMed but intentionally not
    emitted. Records carrying only MedlineDate (no Year) yield ``""``.
    """
    if pubdate is None:
        return ""
    year_text = join_text(pubdate.find("Year"))
    if not year_text:
        return ""
    try:
        year = f"{int(year_text):04d}"
    except ValueError:
        return ""
    month = normalize_month(join_text(pubdate.find("Month")))
    if month is not None:
        return f"{year}-{month:02d}"
    return year


def format_author(author):
    """Format an Author element as ``LastName Initials`` (LastInitials)."""
    if author is None:
        return ""
    last = join_text(author.find("LastName"))
    if not last:
        return join_text(author.find("CollectiveName"))
    initials = join_text(author.find("Initials"))
    if initials:
        return f"{last} {initials}"
    return last


def parse_article(elem):
    """Extract one row (dict keyed by COLUMNS) from a PubmedArticle element."""
    article = elem.find("./MedlineCitation/Article")
    journal = article.find("Journal") if article is not None else None
    authors = (
        article.findall("./AuthorList/Author") if article is not None else []
    )
    pub_types = (
        [
            join_text(pt)
            for pt in article.findall("./PublicationTypeList/PublicationType")
        ]
        if article is not None
        else []
    )
    doi = elem.find("./PubmedData/ArticleIdList/ArticleId[@IdType='doi']")
    pub_date_elem = (
        journal.find("./JournalIssue/PubDate") if journal is not None else None
    )
    return {
        "PMID": join_text(elem.find("./MedlineCitation/PMID")),
        "DOI": join_text(doi),
        "Title": join_text(article.find("ArticleTitle"))
        if article is not None
        else "",
        "Journal": join_text(journal.find("Title"))
        if journal is not None
        else "",
        "ISSN": join_text(journal.find("ISSN")) if journal is not None else "",
        "PubDate": normalize_pubdate(pub_date_elem),
        "FirstAuthor": format_author(authors[0]) if authors else "",
        "LastAuthor": format_author(authors[-1]) if authors else "",
        "PublicationTypes": ";".join(pt for pt in pub_types if pt),
    }


def iter_records(stream):
    """Yield (tag, element) for top-level records, clearing processed memory.

    Uses ElementTree iterparse; after each yielded record is consumed, the
    element is cleared and detached from the root so memory stays bounded.
    """
    root = None
    for event, elem in ET.iterparse(stream, events=("start", "end")):
        if event == "start":
            if root is None:
                root = elem
            continue
        if elem.tag in ("PubmedArticle", "DeleteCitation"):
            yield elem.tag, elem
            elem.clear()
            if root is not None:
                try:
                    root.remove(elem)
                except ValueError:
                    pass


def parse_files(files, output, summary_json=None):
    """Parse updatefiles into one workbook; return and optionally dump summary.

    Args:
        files: Input ``.xml`` / ``.xml.gz`` paths (order preserved in output).
        output: Destination ``.xlsx`` path.
        summary_json: Optional path for the JSON summary.

    Returns:
        Dict: {source_files, article_count, deleted_pmids, first_rows}.
    """
    rows = []
    deleted_pmids = []

    for path in files:
        with open_maybe_gzip(path) as stream:
            for tag, elem in iter_records(stream):
                if tag == "PubmedArticle":
                    rows.append(parse_article(elem))
                else:
                    deleted_pmids.extend(
                        join_text(p) for p in elem.findall("./PMID")
                    )

    workbook = Workbook(write_only=True)

    articles_sheet = workbook.create_sheet(ARTICLE_SHEET)
    articles_sheet.append(COLUMNS)
    for row in rows:
        articles_sheet.append([row[column] for column in COLUMNS])

    deleted_sheet = workbook.create_sheet(DELETED_SHEET)
    deleted_sheet.append(["PMID"])
    for pmid in deleted_pmids:
        deleted_sheet.append([pmid])

    workbook.save(output)

    summary = {
        "source_files": [os.path.basename(path) for path in files],
        "article_count": len(rows),
        "deleted_pmids": deleted_pmids,
        "first_rows": rows[:3],
    }
    if summary_json:
        with open(summary_json, "w", encoding="utf-8") as handle:
            json.dump(summary, handle, ensure_ascii=False, indent=2)
    return summary


def main(argv=None):
    """Command-line entry point."""
    parser = argparse.ArgumentParser(
        description=(
            "Streaming parser for PubMed updatefiles XML(.gz) into one "
            "Excel workbook (PubMed Articles + Deleted PMIDs sheets)."
        )
    )
    parser.add_argument(
        "files",
        nargs="+",
        help="Input PubMed updatefiles (.xml or .xml.gz), order preserved",
    )
    parser.add_argument(
        "-o",
        "--output",
        required=True,
        help="Output .xlsx path (e.g. combined.xlsx)",
    )
    parser.add_argument(
        "--summary-json",
        dest="summary_json",
        default=None,
        help="Optional path to write the JSON summary",
    )
    args = parser.parse_args(argv)

    try:
        summary = parse_files(args.files, args.output, args.summary_json)
    except FileNotFoundError as exc:
        print(f"Error: input file not found: {exc}", file=sys.stderr)
        return 1

    print(
        "Parsed {} article(s) and {} deleted PMID(s) from {} file(s)".format(
            summary["article_count"],
            len(summary["deleted_pmids"]),
            len(summary["source_files"]),
        )
    )
    print(f"Output: {args.output}")
    if args.summary_json:
        print(f"Summary: {args.summary_json}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
